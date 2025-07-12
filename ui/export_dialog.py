
from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QPushButton,
                             QLabel, QGroupBox, QRadioButton, QCheckBox, QComboBox,
                             QFileDialog, QProgressBar, QTextEdit, QMessageBox,
                             QButtonGroup, QFrame, QGridLayout)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from database.mysql import mysql_manager
from database.user_manager import user_manager
import json
import csv
from datetime import datetime
import os

class ExportWorkerThread(QThread):
    """Dışa aktarma işlemi için worker thread"""
    progress = pyqtSignal(int)
    status = pyqtSignal(str)
    finished = pyqtSignal(str)
    error = pyqtSignal(str)
    
    def __init__(self, export_type, file_path, options):
        super().__init__()
        self.export_type = export_type
        self.file_path = file_path
        self.options = options
        
    def run(self):
        try:
            if self.export_type == "categories":
                self.export_categories()
            elif self.export_type == "account_categories":
                self.export_account_categories()
            elif self.export_type == "statistics":
                self.export_statistics()
            elif self.export_type == "excel":
                self.export_to_excel()
                
        except Exception as e:
            self.error.emit(str(e))
            
    def export_categories(self):
        """Kategorileri dışa aktar"""
        self.status.emit("Kategoriler yükleniyor...")
        categories = mysql_manager.get_categories('icerik')
        
        self.progress.emit(25)
        
        if self.file_path.endswith('.json'):
            export_data = {
                'export_date': datetime.now().isoformat(),
                'export_type': 'categories',
                'categories': categories
            }
            
            with open(self.file_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)
                
        elif self.file_path.endswith('.txt'):
            with open(self.file_path, 'w', encoding='utf-8') as f:
                f.write("# Kategori Listesi\n")
                f.write(f"# Dışa aktarma tarihi: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("# Format: kategori_turu:ana_kategori:alt_kategori:aciklama\n\n")
                
                for cat in categories:
                    line = f"icerik:{cat.get('ana_kategori', '')}:{cat.get('alt_kategori', '')}:{cat.get('aciklama', '')}\n"
                    f.write(line)
                    
        elif self.file_path.endswith('.csv'):
            with open(self.file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Kategori Türü', 'Ana Kategori', 'Alt Kategori', 'Açıklama'])
                
                for cat in categories:
                    writer.writerow([
                        'icerik',
                        cat.get('ana_kategori', ''),
                        cat.get('alt_kategori', ''),
                        cat.get('aciklama', '')
                    ])
        
        self.progress.emit(100)
        self.finished.emit("Kategoriler başarıyla dışa aktarıldı")
        
    def export_account_categories(self):
        """Hesap kategorilerini dışa aktar"""
        self.status.emit("Hesap kategorileri yükleniyor...")
        
        include_login = self.options.get('include_login', True)
        include_target = self.options.get('include_target', True)
        
        all_data = []
        
        # Giriş yapılan hesaplar
        if include_login:
            self.status.emit("Giriş yapılan hesaplar işleniyor...")
            login_accounts = user_manager.get_all_users()
            for i, acc in enumerate(login_accounts):
                categories = mysql_manager.get_account_categories(acc['kullanici_adi'], 'giris_yapilan')
                for cat in categories:
                    all_data.append({
                        'kullanici_adi': acc['kullanici_adi'],
                        'hesap_turu': 'giris_yapilan',
                        'ana_kategori': cat.get('ana_kategori', ''),
                        'alt_kategori': cat.get('alt_kategori', ''),
                        'kategori_degeri': cat.get('kategori_degeri', '')
                    })
                self.progress.emit(int((i + 1) / len(login_accounts) * 50))
        
        # Hedef hesaplar
        if include_target:
            self.status.emit("Hedef hesaplar işleniyor...")
            target_accounts = mysql_manager.get_all_targets()
            for i, acc in enumerate(target_accounts):
                categories = mysql_manager.get_account_categories(acc['kullanici_adi'], 'hedef')
                for cat in categories:
                    all_data.append({
                        'kullanici_adi': acc['kullanici_adi'],
                        'hesap_turu': 'hedef',
                        'ana_kategori': cat.get('ana_kategori', ''),
                        'alt_kategori': cat.get('alt_kategori', ''),
                        'kategori_degeri': cat.get('kategori_degeri', '')
                    })
                progress_val = 50 + int((i + 1) / len(target_accounts) * 50)
                self.progress.emit(progress_val)
        
        # Dosyaya kaydet
        self.status.emit("Dosyaya kaydediliyor...")
        
        if self.file_path.endswith('.json'):
            export_data = {
                'export_date': datetime.now().isoformat(),
                'export_type': 'account_categories',
                'data': all_data
            }
            
            with open(self.file_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)
                
        elif self.file_path.endswith('.txt'):
            with open(self.file_path, 'w', encoding='utf-8') as f:
                f.write("# Hesap Kategorileri\n")
                f.write(f"# Dışa aktarma tarihi: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("# Format: kullanici_adi:hesap_turu:ana_kategori:alt_kategori:kategori_degeri\n\n")
                
                for data in all_data:
                    line = f"{data['kullanici_adi']}:{data['hesap_turu']}:{data['ana_kategori']}:{data['alt_kategori']}:{data['kategori_degeri']}\n"
                    f.write(line)
                    
        elif self.file_path.endswith('.csv'):
            with open(self.file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Kullanıcı Adı', 'Hesap Türü', 'Ana Kategori', 'Alt Kategori', 'Kategori Değeri'])
                
                for data in all_data:
                    writer.writerow([
                        data['kullanici_adi'],
                        data['hesap_turu'],
                        data['ana_kategori'],
                        data['alt_kategori'],
                        data['kategori_degeri']
                    ])
        
        self.progress.emit(100)
        self.finished.emit(f"Hesap kategorileri başarıyla dışa aktarıldı ({len(all_data)} kayıt)")
        
    def export_to_excel(self):
        """Excel formatında dışa aktar"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            
            # Kategoriler sayfası
            ws_categories = wb.active
            ws_categories.title = "Kategoriler"
            
            # Başlıklar
            headers = ['Ana Kategori', 'Alt Kategori', 'Açıklama', 'Oluşturma Tarihi']
            for col, header in enumerate(headers, 1):
                cell = ws_categories.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Kategori verileri
            categories = mysql_manager.get_categories('icerik')
            for row, cat in enumerate(categories, 2):
                ws_categories.cell(row=row, column=1, value=cat.get('ana_kategori', ''))
                ws_categories.cell(row=row, column=2, value=cat.get('alt_kategori', ''))
                ws_categories.cell(row=row, column=3, value=cat.get('aciklama', ''))
                ws_categories.cell(row=row, column=4, value=cat.get('created_at', ''))
                
                self.progress.emit(int(row / len(categories) * 50))
            
            # Hesap kategorileri sayfası
            ws_accounts = wb.create_sheet("Hesap Kategorileri")
            
            headers = ['Kullanıcı Adı', 'Hesap Türü', 'Ana Kategori', 'Alt Kategori', 'Kategori Değeri']
            for col, header in enumerate(headers, 1):
                cell = ws_accounts.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Hesap verileri
            row_num = 2
            
            # Giriş yapılan hesaplar
            if self.options.get('include_login', True):
                login_accounts = user_manager.get_all_users()
                for acc in login_accounts:
                    categories = mysql_manager.get_account_categories(acc['kullanici_adi'], 'giris_yapilan')
                    for cat in categories:
                        ws_accounts.cell(row=row_num, column=1, value=acc['kullanici_adi'])
                        ws_accounts.cell(row=row_num, column=2, value='Giriş Yapılan')
                        ws_accounts.cell(row=row_num, column=3, value=cat.get('ana_kategori', ''))
                        ws_accounts.cell(row=row_num, column=4, value=cat.get('alt_kategori', ''))
                        ws_accounts.cell(row=row_num, column=5, value=cat.get('kategori_degeri', ''))
                        row_num += 1
            
            # Hedef hesaplar
            if self.options.get('include_target', True):
                target_accounts = mysql_manager.get_all_targets()
                for acc in target_accounts:
                    categories = mysql_manager.get_account_categories(acc['kullanici_adi'], 'hedef')
                    for cat in categories:
                        ws_accounts.cell(row=row_num, column=1, value=acc['kullanici_adi'])
                        ws_accounts.cell(row=row_num, column=2, value='Hedef')
                        ws_accounts.cell(row=row_num, column=3, value=cat.get('ana_kategori', ''))
                        ws_accounts.cell(row=row_num, column=4, value=cat.get('alt_kategori', ''))
                        ws_accounts.cell(row=row_num, column=5, value=cat.get('kategori_degeri', ''))
                        row_num += 1
            
            # İstatistikler sayfası
            ws_stats = wb.create_sheet("İstatistikler")
            
            # Genel istatistikler
            stats_data = [
                ['Metrik', 'Değer'],
                ['Toplam Kategori Sayısı', len(set(cat.get('ana_kategori', '') for cat in categories))],
                ['Toplam Alt Kategori Sayısı', len([cat for cat in categories if cat.get('alt_kategori')])],
                ['Toplam Giriş Yapılan Hesap', len(user_manager.get_all_users())],
                ['Toplam Hedef Hesap', len(mysql_manager.get_all_targets())],
                ['Dışa Aktarma Tarihi', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
            ]
            
            for row, (metric, value) in enumerate(stats_data, 1):
                cell1 = ws_stats.cell(row=row, column=1, value=metric)
                cell2 = ws_stats.cell(row=row, column=2, value=value)
                
                if row == 1:  # Başlık
                    cell1.font = Font(bold=True)
                    cell2.font = Font(bold=True)
                    cell1.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
                    cell2.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
            
            # Sütun genişliklerini ayarla
            for ws in wb.worksheets:
                for column_cells in ws.columns:
                    length = max(len(str(cell.value or '')) for cell in column_cells)
                    ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
            
            # Dosyayı kaydet
            wb.save(self.file_path)
            
            self.progress.emit(100)
            self.finished.emit("Excel dosyası başarıyla oluşturuldu")
            
        except ImportError:
            self.error.emit("Excel dışa aktarma için openpyxl kütüphanesi gerekli. Lütfen 'pip install openpyxl' çalıştırın.")
        except Exception as e:
            self.error.emit(f"Excel dışa aktarma hatası: {str(e)}")

class ExportDialog(QDialog):
    """Dışa aktarma dialog'u"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📤 Kategori Dışa Aktarma")
        self.setModal(True)
        self.resize(600, 500)
        
        self.worker_thread = None
        self.setup_ui()
        
    def setup_ui(self):
        """UI'yi ayarla"""
        layout = QVBoxLayout()
        
        # Başlık
        title_label = QLabel("📤 Kategori Verilerini Dışa Aktarma")
        title_label.setStyleSheet("""
            font-size: 20px;
            font-weight: 700;
            color: #1e293b;
            margin-bottom: 15px;
        """)
        layout.addWidget(title_label)
        
        # Dışa aktarma türü
        export_type_group = QGroupBox("Dışa Aktarılacak Veri Türü")
        export_type_layout = QVBoxLayout()
        
        self.export_type_group = QButtonGroup()
        
        self.categories_radio = QRadioButton("📋 Sadece Kategoriler")
        self.categories_radio.setChecked(True)
        self.export_type_group.addButton(self.categories_radio, 0)
        
        self.account_categories_radio = QRadioButton("👥 Hesap Kategorileri")
        self.export_type_group.addButton(self.account_categories_radio, 1)
        
        self.full_export_radio = QRadioButton("📊 Tam Dışa Aktarma (Kategoriler + Hesaplar + İstatistikler)")
        self.export_type_group.addButton(self.full_export_radio, 2)
        
        export_type_layout.addWidget(self.categories_radio)
        export_type_layout.addWidget(self.account_categories_radio)
        export_type_layout.addWidget(self.full_export_radio)
        export_type_group.setLayout(export_type_layout)
        
        # Hesap türü seçimi
        self.account_type_group = QGroupBox("Dahil Edilecek Hesap Türleri")
        account_type_layout = QVBoxLayout()
        
        self.include_login_checkbox = QCheckBox("🔐 Giriş Yapılan Hesaplar")
        self.include_login_checkbox.setChecked(True)
        
        self.include_target_checkbox = QCheckBox("🎯 Hedef Hesaplar")
        self.include_target_checkbox.setChecked(True)
        
        account_type_layout.addWidget(self.include_login_checkbox)
        account_type_layout.addWidget(self.include_target_checkbox)
        self.account_type_group.setLayout(account_type_layout)
        
        # Dosya formatı
        format_group = QGroupBox("Dosya Formatı")
        format_layout = QVBoxLayout()
        
        self.format_combo = QComboBox()
        self.format_combo.addItems([
            "📄 JSON (.json)",
            "📝 Metin Dosyası (.txt)",
            "📊 CSV (.csv)",
            "📗 Excel (.xlsx)"
        ])
        
        format_layout.addWidget(self.format_combo)
        format_group.setLayout(format_layout)
        
        # İlerleme çubuğu ve durum
        progress_group = QGroupBox("İşlem Durumu")
        progress_layout = QVBoxLayout()
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        
        self.status_label = QLabel("Dışa aktarma için ayarları yapın ve başlatın")
        self.status_label.setStyleSheet("color: #64748b; font-size: 13px;")
        
        progress_layout.addWidget(self.progress_bar)
        progress_layout.addWidget(self.status_label)
        progress_group.setLayout(progress_layout)
        
        # Butonlar
        button_layout = QHBoxLayout()
        
        self.export_btn = QPushButton("📤 Dışa Aktarma Başlat")
        self.export_btn.setStyleSheet("""
            QPushButton {
                background: #3b82f6;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 20px;
                font-weight: 600;
            }
            QPushButton:hover {
                background: #2563eb;
            }
        """)
        self.export_btn.clicked.connect(self.start_export)
        
        self.close_btn = QPushButton("❌ Kapat")
        self.close_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.export_btn)
        button_layout.addStretch()
        button_layout.addWidget(self.close_btn)
        
        # Layout'a ekle
        layout.addWidget(export_type_group)
        layout.addWidget(self.account_type_group)
        layout.addWidget(format_group)
        layout.addWidget(progress_group)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
        
        # Signal bağlantıları
        self.export_type_group.buttonToggled.connect(self.on_export_type_changed)
        self.on_export_type_changed()  # İlk durumu ayarla
        
    def on_export_type_changed(self):
        """Dışa aktarma türü değiştiğinde"""
        export_type = self.export_type_group.checkedId()
        
        # Hesap türü seçimi sadece hesap kategorileri için gerekli
        self.account_type_group.setEnabled(export_type in [1, 2])
        
    def start_export(self):
        """Dışa aktarmayı başlat"""
        try:
            # Dosya yolu seç
            export_type = self.export_type_group.checkedId()
            format_text = self.format_combo.currentText()
            
            if "JSON" in format_text:
                extension = "json"
                filter_text = "JSON Dosyaları (*.json)"
            elif "Metin" in format_text:
                extension = "txt"
                filter_text = "Metin Dosyaları (*.txt)"
            elif "CSV" in format_text:
                extension = "csv"
                filter_text = "CSV Dosyaları (*.csv)"
            elif "Excel" in format_text:
                extension = "xlsx"
                filter_text = "Excel Dosyaları (*.xlsx)"
            
            # Varsayılan dosya adı
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            if export_type == 0:
                default_name = f"kategoriler_{timestamp}.{extension}"
            elif export_type == 1:
                default_name = f"hesap_kategorileri_{timestamp}.{extension}"
            else:
                default_name = f"tam_disa_aktarma_{timestamp}.{extension}"
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Dışa Aktarma Dosyası Seç", 
                default_name, filter_text
            )
            
            if not file_path:
                return
            
            # Seçenekleri hazırla
            options = {
                'include_login': self.include_login_checkbox.isChecked(),
                'include_target': self.include_target_checkbox.isChecked()
            }
            
            # Worker thread'i başlat
            if extension == "xlsx":
                thread_export_type = "excel"
            elif export_type == 0:
                thread_export_type = "categories"
            elif export_type == 1:
                thread_export_type = "account_categories"
            else:
                thread_export_type = "statistics"
            
            self.worker_thread = ExportWorkerThread(thread_export_type, file_path, options)
            self.worker_thread.progress.connect(self.progress_bar.setValue)
            self.worker_thread.status.connect(self.status_label.setText)
            self.worker_thread.finished.connect(self.on_export_finished)
            self.worker_thread.error.connect(self.on_export_error)
            
            # UI'yi güncelle
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(0)
            self.export_btn.setEnabled(False)
            self.export_btn.setText("⏳ Dışa Aktarılıyor...")
            
            self.worker_thread.start()
            
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Dışa aktarma başlatılamadı: {str(e)}")
            
    def on_export_finished(self, message):
        """Dışa aktarma tamamlandığında"""
        self.progress_bar.setVisible(False)
        self.export_btn.setEnabled(True)
        self.export_btn.setText("📤 Dışa Aktarma Başlat")
        self.status_label.setText("✅ " + message)
        
        QMessageBox.information(self, "Başarılı", message)
        
    def on_export_error(self, error_message):
        """Dışa aktarma hatası"""
        self.progress_bar.setVisible(False)
        self.export_btn.setEnabled(True)
        self.export_btn.setText("📤 Dışa Aktarma Başlat")
        self.status_label.setText("❌ " + error_message)
        
        QMessageBox.critical(self, "Hata", error_message)
        
    def closeEvent(self, event):
        """Dialog kapatılırken"""
        if self.worker_thread and self.worker_thread.isRunning():
            self.worker_thread.terminate()
            self.worker_thread.wait()
        event.accept()
