from PyQt5.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QPushButton,
                             QLabel, QGroupBox, QTabWidget, QWidget, QFileDialog,
                             QTextEdit, QProgressBar, QMessageBox, QCheckBox,
                             QComboBox, QSpinBox, QFormLayout, QFrame, QSplitter,
                             QListWidget, QListWidgetItem, QTableWidget, QTableWidgetItem,
                             QHeaderView, QScrollArea, QGridLayout, QSpacerItem, QSizePolicy)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QPropertyAnimation, QEasingCurve
from PyQt5.QtGui import QFont, QIcon, QLinearGradient, QColor, QPalette
from database.mysql import mysql_manager
from database.user_manager import user_manager
import json
import csv
import os
from datetime import datetime
import openpyxl

class ModernProgressBar(QProgressBar):
    """Modern progress bar"""
    def __init__(self):
        super().__init__()
        self.setStyleSheet("""
            QProgressBar {
                border: 2px solid #e2e8f0;
                border-radius: 8px;
                background: #f8fafc;
                text-align: center;
                font-weight: 600;
                font-size: 13px;
                color: #374151;
            }
            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3b82f6, stop:1 #2563eb);
                border-radius: 6px;
            }
        """)
        self.setMinimum(0)
        self.setMaximum(100)

class AdvancedImportThread(QThread):
    """Gelişmiş içe aktarma thread'i"""
    progress = pyqtSignal(int)
    status = pyqtSignal(str)
    log = pyqtSignal(str)
    finished = pyqtSignal(str)
    error = pyqtSignal(str)

    def __init__(self, operation_type, file_paths, options):
        super().__init__()
        self.operation_type = operation_type
        self.file_paths = file_paths
        self.options = options

    def run(self):
        try:
            if self.operation_type == "batch_import":
                self.batch_import_files()
            elif self.operation_type == "excel_import":
                self.import_from_excel()
            elif self.operation_type == "merge_categories":
                self.merge_category_files()
            elif self.operation_type == "validate_data":
                self.validate_import_data()
        except Exception as e:
            self.error.emit(str(e))

    def batch_import_files(self):
        """Toplu dosya içe aktarma"""
        total_files = len(self.file_paths)
        imported_categories = 0
        imported_accounts = 0

        for i, file_path in enumerate(self.file_paths):
            self.status.emit(f"İşleniyor: {os.path.basename(file_path)}")

            try:
                if file_path.endswith('.json'):
                    result = self.import_json_file(file_path)
                elif file_path.endswith('.txt'):
                    result = self.import_txt_file(file_path)
                elif file_path.endswith('.csv'):
                    result = self.import_csv_file(file_path)
                else:
                    self.log.emit(f"❌ Desteklenmeyen dosya formatı: {file_path}")
                    continue

                imported_categories += result.get('categories', 0)
                imported_accounts += result.get('accounts', 0)

                self.log.emit(f"✅ {os.path.basename(file_path)}: {result.get('categories', 0)} kategori, {result.get('accounts', 0)} hesap")

            except Exception as e:
                self.log.emit(f"❌ {os.path.basename(file_path)}: {str(e)}")

            self.progress.emit(int((i + 1) / total_files * 100))

        self.finished.emit(f"✅ Toplu içe aktarma tamamlandı: {imported_categories} kategori, {imported_accounts} hesap")

    def import_json_file(self, file_path):
        """JSON dosyası içe aktarma"""
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        categories_count = 0
        accounts_count = 0

        # Profil kategorileri
        if 'profil_kategorileri' in data:
            for cat in data['profil_kategorileri']:
                if mysql_manager.add_hierarchical_category(
                    cat.get('kategori_turu', 'profil'),
                    cat.get('ana_kategori', ''),
                    None,
                    cat.get('aciklama', '')
                ):
                    categories_count += 1

        # İçerik kategorileri
        if 'icerik_kategorileri' in data:
            for cat in data['icerik_kategorileri']:
                kategori_turu = cat.get('kategori_turu', 'icerik')
                ana_kategori = cat.get('ana_kategori', '')
                alt_kategoriler = cat.get('alt_kategoriler')
                aciklama = cat.get('aciklama', '')
                
                # Ana kategoriyi ekle
                if mysql_manager.add_hierarchical_category(
                    kategori_turu,
                    ana_kategori,
                    None,
                    aciklama
                ):
                    categories_count += 1
                
                # Alt kategorileri güncelle
                if alt_kategoriler:
                    connection = mysql_manager.get_connection()
                    if connection:
                        try:
                            cursor = connection.cursor()
                            update_query = """
                            UPDATE kategoriler 
                            SET alt_kategoriler = %s 
                            WHERE kategori_turu = %s AND ana_kategori = %s
                            """
                            cursor.execute(update_query, (alt_kategoriler, kategori_turu, ana_kategori))
                            connection.commit()
                        except Exception as e:
                            self.log.emit(f"❌ Alt kategori güncelleme hatası: {str(e)}")
                        finally:
                            if connection.is_connected():
                                cursor.close()
                                connection.close()

        # Eski format kategoriler (geriye uyumluluk)
        if 'categories' in data:
            for cat in data['categories']:
                if mysql_manager.add_hierarchical_category(
                    'icerik', 
                    cat.get('ana_kategori', ''),
                    cat.get('alt_kategori'),
                    cat.get('aciklama', '')
                ):
                    categories_count += 1

        # Hesap kategorileri
        if 'account_categories' in data or 'data' in data:
            account_data = data.get('account_categories', data.get('data', []))
            for acc_cat in account_data:
                if mysql_manager.assign_hierarchical_category_to_account(
                    acc_cat.get('kullanici_adi', ''),
                    acc_cat.get('hesap_turu', 'hedef'),
                    acc_cat.get('ana_kategori', ''),
                    acc_cat.get('alt_kategori'),
                    acc_cat.get('kategori_degeri', 'İçe Aktarıldı')
                ):
                    accounts_count += 1

        return {'categories': categories_count, 'accounts': accounts_count}

    def import_txt_file(self, file_path):
        """TXT dosyası içe aktarma"""
        categories_count = 0
        accounts_count = 0

        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue

                parts = line.split(':')
                if len(parts) >= 3:
                    if parts[0] == 'icerik':  # Kategori
                        if mysql_manager.add_hierarchical_category(
                            'icerik', parts[1], 
                            parts[2] if parts[2] else None,
                            parts[3] if len(parts) > 3 else ''
                        ):
                            categories_count += 1
                    else:  # Hesap kategorisi
                        if len(parts) >= 5:
                            if mysql_manager.assign_hierarchical_category_to_account(
                                parts[0], parts[1], parts[2], 
                                parts[3] if parts[3] else None,
                                parts[4] if len(parts) > 4 else 'İçe Aktarıldı'
                            ):
                                accounts_count += 1

        return {'categories': categories_count, 'accounts': accounts_count}

    def import_csv_file(self, file_path):
        """CSV dosyası içe aktarma"""
        categories_count = 0
        accounts_count = 0

        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            header = next(reader, None)

            if not header:
                return {'categories': 0, 'accounts': 0}

            # Kategori dosyası mı hesap dosyası mı kontrol et
            if 'Ana Kategori' in header and 'Kullanıcı Adı' not in header:
                # Kategori dosyası
                for row in reader:
                    if len(row) >= 2:
                        if mysql_manager.add_hierarchical_category(
                            'icerik', row[0], 
                            row[1] if row[1] else None,
                            row[2] if len(row) > 2 else ''
                        ):
                            categories_count += 1

            elif 'Kullanıcı Adı' in header:
                # Hesap kategorisi dosyası
                for row in reader:
                    if len(row) >= 4:
                        if mysql_manager.assign_hierarchical_category_to_account(
                            row[0], row[1], row[2],
                            row[3] if row[3] else None,
                            row[4] if len(row) > 4 else 'İçe Aktarıldı'
                        ):
                            accounts_count += 1

        return {'categories': categories_count, 'accounts': accounts_count}

    def import_from_excel(self):
        """Excel dosyasından içe aktarma"""
        wb = openpyxl.load_workbook(self.file_paths[0])
        imported_categories = 0
        imported_accounts = 0

        # Kategoriler sayfası
        if 'Kategoriler' in wb.sheetnames:
            ws = wb['Kategoriler']
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Ana kategori varsa
                    if mysql_manager.add_hierarchical_category(
                        'icerik', row[0], 
                        row[1] if row[1] else None,
                        row[2] if len(row) > 2 and row[2] else ''
                    ):
                        imported_categories += 1

        # Hesap kategorileri sayfası
        if 'Hesap Kategorileri' in wb.sheetnames:
            ws = wb['Hesap Kategorileri']
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[2]:  # Kullanıcı adı ve ana kategori varsa
                    hesap_turu = 'giris_yapilan' if row[1] == 'Giriş Yapılan' else 'hedef'
                    if mysql_manager.assign_hierarchical_category_to_account(
                        row[0], hesap_turu, row[2],
                        row[3] if len(row) > 3 and row[3] else None,
                        row[4] if len(row) > 4 and row[4] else 'İçe Aktarıldı'
                    ):
                        imported_accounts += 1

        self.finished.emit(f"✅ Excel içe aktarma tamamlandı: {imported_categories} kategori, {imported_accounts} hesap")

    def validate_import_data(self):
        """İçe aktarma verilerini doğrula"""
        validation_results = []

        for file_path in self.file_paths:
            self.status.emit(f"Doğrulanıyor: {os.path.basename(file_path)}")

            try:
                result = self.validate_file(file_path)
                validation_results.append({
                    'file': os.path.basename(file_path),
                    'valid_lines': result['valid'],
                    'invalid_lines': result['invalid'],
                    'errors': result['errors']
                })

            except Exception as e:
                validation_results.append({
                    'file': os.path.basename(file_path),
                    'valid_lines': 0,
                    'invalid_lines': 0,
                    'errors': [str(e)]
                })

        # Sonuçları logla
        for result in validation_results:
            self.log.emit(f"📁 {result['file']}")
            self.log.emit(f"  ✅ Geçerli: {result['valid_lines']}")
            self.log.emit(f"  ❌ Geçersiz: {result['invalid_lines']}")
            for error in result['errors']:
                self.log.emit(f"  🚨 {error}")
            self.log.emit("")

        self.finished.emit("✅ Doğrulama tamamlandı")

    def validate_file(self, file, path):
        """Dosyayı doğrula"""
        valid_count = 0
        invalid_count = 0
        errors = []

        if file.path.endswith('.json'):
            try:
                with open(file.path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                valid_count = 1
            except json.JSONDecodeError as e:
                errors.append(f"JSON formatı geçersiz: {str(e)}")
                invalid_count = 1

        elif file.path.endswith('.txt'):
            with open(file.path, 'r', encoding='utf-8') as f:
                for line_num, line in enumerate(f, 1):
                    line = line.strip()
                    if not line or line.startswith('#'):
                        continue

                    parts = line.split(':')
                    if len(parts) < 3:
                        errors.append(f"Satır {line_num}: Eksik alan")
                        invalid_count += 1
                    else:
                        valid_count += 1

        return {'valid': valid_count, 'invalid': invalid_count, 'errors': errors}

class AdvancedFileOperationsDialog(QDialog):
    """Gelişmiş dosya işlemleri dialog'u - Modern tasarım"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("🔄 Gelişmiş İçe/Dışa Aktarma İşlemleri")
        self.setModal(True)
        self.resize(1200, 800)

        self.worker_thread = None
        self.setup_modern_ui()

    def setup_modern_ui(self):
        """Modern UI'yi ayarla"""
        # Ana stil
        self.setStyleSheet("""
            QDialog {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #f8fafc, stop:1 #e2e8f0);
                border: none;
            }
            QTabWidget::pane {
                border: none;
                background: white;
                border-radius: 16px;
                margin-top: 10px;
            }
            QTabBar::tab {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #f1f5f9, stop:1 #e2e8f0);
                padding: 16px 28px;
                margin-right: 4px;
                border-top-left-radius: 12px;
                border-top-right-radius: 12px;
                border: none;
                font-weight: 600;
                font-size: 14px;
                color: #64748b;
            }
            QTabBar::tab:selected {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #3b82f6, stop:1 #2563eb);
                color: white;
            }
            QTabBar::tab:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #e2e8f0, stop:1 #cbd5e1);
            }
            QGroupBox {
                font-size: 16px;
                font-weight: 600;
                color: #374151;
                border: 2px solid #e2e8f0;
                border-radius: 16px;
                margin-top: 16px;
                padding-top: 16px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 24px;
                padding: 0 12px;
                background: white;
                border-radius: 8px;
            }
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #3b82f6, stop:1 #2563eb);
                color: white;
                border: none;
                border-radius: 12px;
                padding: 14px 28px;
                font-weight: 600;
                font-size: 14px;
                min-width: 120px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #2563eb, stop:1 #1d4ed8);
            }
            QPushButton:pressed {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #1d4ed8, stop:1 #1e40af);
            }
            QListWidget {
                border: 2px solid #e2e8f0;
                border-radius: 12px;
                background: white;
                padding: 8px;
                font-size: 13px;
            }
            QListWidget::item {
                padding: 8px 12px;
                border-radius: 8px;
                margin: 2px 0px;
            }
            QListWidget::item:selected {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #3b82f6, stop:1 #2563eb);
                color: white;
            }
            QTextEdit {
                border: 2px solid #e2e8f0;
                border-radius: 12px;
                background: white;
                padding: 12px;
                font-size: 13px;
                font-family: 'Consolas', 'Monaco', monospace;
            }
            QCheckBox {
                font-size: 14px;
                color: #374151;
                spacing: 10px;
            }
            QCheckBox::indicator {
                width: 20px;
                height: 20px;
                border: 2px solid #d1d5db;
                border-radius: 6px;
                background: white;
            }
            QCheckBox::indicator:checked {
                background: #3b82f6;
                border-color: #3b82f6;
            }
            QLabel {
                color: #374151;
                font-size: 14px;
            }
        """)

        layout = QVBoxLayout()
        layout.setSpacing(24)
        layout.setContentsMargins(32, 32, 32, 32)

        # Modern başlık
        header_frame = QFrame()
        header_layout = QVBoxLayout()

        title_label = QLabel("🔄 Gelişmiş İçe/Dışa Aktarma")
        title_label.setStyleSheet("""
            font-size: 36px;
            font-weight: 800;
            color: #1e293b;
            margin-bottom: 8px;
        """)

        subtitle_label = QLabel("Toplu işlemler ve gelişmiş dosya yönetimi")
        subtitle_label.setStyleSheet("""
            font-size: 18px;
            color: #64748b;
            font-weight: 500;
            margin-bottom: 20px;
        """)

        header_layout.addWidget(title_label)
        header_layout.addWidget(subtitle_label)
        header_frame.setLayout(header_layout)

        layout.addWidget(header_frame)

        # Tab widget
        self.tabs = QTabWidget()

        # Sekmeleri oluştur
        self.batch_import_tab = self.create_batch_import_tab()
        self.tabs.addTab(self.batch_import_tab, "📥 Toplu İçe Aktarma")

        self.excel_tab = self.create_excel_tab()
        self.tabs.addTab(self.excel_tab, "📗 Excel İşlemleri")

        self.validation_tab = self.create_validation_tab()
        self.tabs.addTab(self.validation_tab, "✅ Veri Doğrulama")

        self.backup_tab = self.create_backup_tab()
        self.tabs.addTab(self.backup_tab, "💾 Yedekleme")

        layout.addWidget(self.tabs, 1)

        # İlerleme ve log alanı
        progress_frame = QFrame()
        progress_frame.setStyleSheet("""
            QFrame {
                background: white;
                border: 2px solid #e2e8f0;
                border-radius: 16px;
                padding: 20px;
            }
        """)

        progress_layout = QVBoxLayout()

        status_layout = QHBoxLayout()
        status_icon = QLabel("📊")
        status_icon.setStyleSheet("font-size: 18px;")

        self.status_label = QLabel("Hazır")
        self.status_label.setStyleSheet("font-size: 14px; font-weight: 600; color: #374151;")

        status_layout.addWidget(status_icon)
        status_layout.addWidget(self.status_label)
        status_layout.addStretch()

        self.progress_bar = ModernProgressBar()
        self.progress_bar.setVisible(False)

        self.log_text = QTextEdit()
        self.log_text.setMaximumHeight(180)
        self.log_text.setPlaceholderText("İşlem logları burada görünecek...")

        progress_layout.addLayout(status_layout)
        progress_layout.addWidget(self.progress_bar)
        progress_layout.addWidget(self.log_text)

        progress_frame.setLayout(progress_layout)
        layout.addWidget(progress_frame)

        # Alt butonlar
        button_layout = QHBoxLayout()

        self.clear_log_btn = QPushButton("🗑️ Logları Temizle")
        self.clear_log_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #f59e0b, stop:1 #d97706);
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #d97706, stop:1 #b45309);
            }
        """)
        self.clear_log_btn.clicked.connect(self.log_text.clear)

        self.close_btn = QPushButton("✅ Kapat")
        self.close_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #10b981, stop:1 #059669);
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #059669, stop:1 #047857);
            }
        """)
        self.close_btn.clicked.connect(self.accept)

        button_layout.addWidget(self.clear_log_btn)
        button_layout.addStretch()
        button_layout.addWidget(self.close_btn)

        layout.addLayout(button_layout)
        self.setLayout(layout)

    def create_batch_import_tab(self):
        """Toplu içe aktarma sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(20)

        # Bilgi kartı
        info_card = QFrame()
        info_card.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #dbeafe, stop:1 #bfdbfe);
                border: 2px solid #93c5fd;
                border-radius: 16px;
                padding: 20px;
            }
        """)

        info_layout = QVBoxLayout()

        info_title = QLabel("📥 Toplu İçe Aktarma")
        info_title.setStyleSheet("font-size: 20px; font-weight: 700; color: #1e40af; margin-bottom: 8px;")

        info_desc = QLabel("Birden fazla dosyayı aynı anda içe aktarın\n• JSON, TXT, CSV formatlarını destekler\n• Kategori ve hesap verilerini otomatik tanır\n• Hata durumunda diğer dosyalar işlenmeye devam eder")
        info_desc.setStyleSheet("color: #3730a3; font-size: 14px; line-height: 1.4;")

        info_layout.addWidget(info_title)
        info_layout.addWidget(info_desc)
        info_card.setLayout(info_layout)

        layout.addWidget(info_card)

        # Dosya seçimi
        file_frame = QGroupBox("📁 Dosya Seçimi")
        file_layout = QVBoxLayout()

        file_buttons_layout = QHBoxLayout()
        self.select_files_btn = QPushButton("📁 Dosyaları Seç")
        self.clear_files_btn = QPushButton("🗑️ Listeyi Temizle")
        self.clear_files_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #ef4444, stop:1 #dc2626);
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #dc2626, stop:1 #b91c1c);
            }
        """)

        self.select_files_btn.clicked.connect(self.select_batch_files)
        self.clear_files_btn.clicked.connect(self.clear_file_list)

        file_buttons_layout.addWidget(self.select_files_btn)
        file_buttons_layout.addWidget(self.clear_files_btn)
        file_buttons_layout.addStretch()

        self.file_list = QListWidget()
        self.file_list.setMinimumHeight(150)

        file_layout.addLayout(file_buttons_layout)
        file_layout.addWidget(self.file_list)
        file_frame.setLayout(file_layout)

        layout.addWidget(file_frame)

        # Seçenekler
        options_frame = QGroupBox("⚙️ İçe Aktarma Seçenekleri")
        options_layout = QVBoxLayout()

        self.skip_duplicates_check = QCheckBox("Mevcut kategorileri atla")
        self.skip_duplicates_check.setChecked(True)

        self.create_backup_check = QCheckBox("İşlem öncesi yedek oluştur")
        self.create_backup_check.setChecked(True)

        options_layout.addWidget(self.skip_duplicates_check)
        options_layout.addWidget(self.create_backup_check)
        options_frame.setLayout(options_layout)

        layout.addWidget(options_frame)

        # İşlem başlat
        self.start_batch_btn = QPushButton("🚀 Toplu İçe Aktarmayı Başlat")
        self.start_batch_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #10b981, stop:1 #059669);
                font-size: 16px;
                padding: 16px 32px;
                min-height: 20px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #059669, stop:1 #047857);
            }
        """)
        self.start_batch_btn.clicked.connect(self.start_batch_import)

        layout.addWidget(self.start_batch_btn)
        layout.addStretch()

        widget.setLayout(layout)
        return widget

    def create_excel_tab(self):
        """Excel işlemleri sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(20)

        # Excel içe aktarma
        import_frame = QGroupBox("📗 Excel Dosyasından İçe Aktarma")
        import_layout = QVBoxLayout()

        excel_info = QFrame()
        excel_info.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #d1fae5, stop:1 #a7f3d0);
                border: 2px solid #6ee7b7;
                border-radius: 12px;
                padding: 16px;
            }
        """)

        excel_info_layout = QVBoxLayout()
        excel_info_label = QLabel("Excel dosyasında aşağıdaki sayfa adları aranır:\n• Kategoriler: Ana Kategori, Alt Kategori, Açıklama\n• Hesap Kategorileri: Kullanıcı Adı, Hesap Türü, Ana Kategori, Alt Kategori, Kategori Değeri")
        excel_info_label.setStyleSheet("color: #065f46; font-size: 13px;")
        excel_info_layout.addWidget(excel_info_label)
        excel_info.setLayout(excel_info_layout)

        excel_buttons_layout = QHBoxLayout()
        self.select_excel_btn = QPushButton("📁 Excel Dosyası Seç")
        self.import_excel_btn = QPushButton("📥 Excel'den İçe Aktar")
        self.import_excel_btn.setEnabled(False)

        self.select_excel_btn.clicked.connect(self.select_excel_file)
        self.import_excel_btn.clicked.connect(self.import_from_excel)

        excel_buttons_layout.addWidget(self.select_excel_btn)
        excel_buttons_layout.addWidget(self.import_excel_btn)
        excel_buttons_layout.addStretch()

        self.excel_file_label = QLabel("Dosya seçilmedi")
        self.excel_file_label.setStyleSheet("color: #64748b; font-style: italic; font-size: 13px;")

        import_layout.addWidget(excel_info)
        import_layout.addLayout(excel_buttons_layout)
        import_layout.addWidget(self.excel_file_label)
        import_frame.setLayout(import_layout)

        layout.addWidget(import_frame)

        # Excel şablonu
        template_frame = QGroupBox("📋 Excel Şablonu")
        template_layout = QVBoxLayout()

        template_info = QLabel("Boş Excel şablonu indirerek kendi verilerinizi hazırlayabilirsiniz")
        template_info.setStyleSheet("color: #64748b; font-size: 14px;")

        self.download_template_btn = QPushButton("⬇️ Excel Şablonunu İndir")
        self.download_template_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #8b5cf6, stop:1 #7c3aed);
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #7c3aed, stop:1 #6d28d9);
            }
        """)
        self.download_template_btn.clicked.connect(self.download_excel_template)

        template_layout.addWidget(template_info)
        template_layout.addWidget(self.download_template_btn)
        template_frame.setLayout(template_layout)

        layout.addWidget(template_frame)
        layout.addStretch()

        widget.setLayout(layout)
        return widget

    def create_validation_tab(self):
        """Veri doğrulama sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(20)

        # Bilgi kartı
        info_card = QFrame()
        info_card.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #fef3c7, stop:1 #fed7aa);
                border: 2px solid #fbbf24;
                border-radius: 16px;
                padding: 20px;
            }
        """)

        info_layout = QVBoxLayout()

        info_title = QLabel("✅ Veri Doğrulama")
        info_title.setStyleSheet("font-size: 20px; font-weight: 700; color: #92400e; margin-bottom: 8px;")

        info_desc = QLabel("İçe aktarma öncesi dosyalarınızı kontrol edin\n• Format hatalarını tespit eder\n• Eksik alanları listeler\n• İçe aktarma öncesi problemleri çözer")
        info_desc.setStyleSheet("color: #78350f; font-size: 14px; line-height: 1.4;")

        info_layout.addWidget(info_title)
        info_layout.addWidget(info_desc)
        info_card.setLayout(info_layout)

        layout.addWidget(info_card)

        # Dosya seçimi
        validation_frame = QGroupBox("🔍 Doğrulanacak Dosyalar")
        validation_layout = QVBoxLayout()

        validation_buttons_layout = QHBoxLayout()
        self.select_validation_files_btn = QPushButton("📁 Dosyaları Seç")
        self.validate_btn = QPushButton("🔍 Doğrulamayı Başlat")
        self.validate_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #f59e0b, stop:1 #d97706);
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #d97706, stop:1 #b45309);
            }
        """)

        self.select_validation_files_btn.clicked.connect(self.select_validation_files)
        self.validate_btn.clicked.connect(self.start_validation)

        validation_buttons_layout.addWidget(self.select_validation_files_btn)
        validation_buttons_layout.addWidget(self.validate_btn)
        validation_buttons_layout.addStretch()

        self.validation_file_list = QListWidget()
        self.validation_file_list.setMinimumHeight(200)

        validation_layout.addLayout(validation_buttons_layout)
        validation_layout.addWidget(self.validation_file_list)
        validation_frame.setLayout(validation_layout)

        layout.addWidget(validation_frame)
        layout.addStretch()

        widget.setLayout(layout)
        return widget

    def create_backup_tab(self):
        """Yedekleme sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(20)

        # Bilgi kartı
        info_card = QFrame()
        info_card.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #e0e7ff, stop:1 #c7d2fe);
                border: 2px solid #a5b4fc;
                border-radius: 16px;
                padding: 20px;
            }
        """)

        info_layout = QVBoxLayout()

        info_title = QLabel("💾 Yedekleme ve Geri Yükleme")
        info_title.setStyleSheet("font-size: 20px; font-weight: 700; color: #3730a3; margin-bottom: 8px;")

        info_desc = QLabel("Verilerinizi güvenle yedekleyin\n• Kategori verilerini yedekle\n• Hesap kategorilerini yedekle\n• Tam veritabanı yedeği oluştur")
        info_desc.setStyleSheet("color: #312e81; font-size: 14px; line-height: 1.4;")

        info_layout.addWidget(info_title)
        info_layout.addWidget(info_desc)
        info_card.setLayout(info_layout)

        layout.addWidget(info_card)

        # Yedekleme seçenekleri
        backup_frame = QGroupBox("💾 Yedekleme Seçenekleri")
        backup_layout = QGridLayout()

        # Butonlar
        self.backup_categories_btn = QPushButton("💾 Kategorileri Yedekle")
        self.backup_categories_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #06b6d4, stop:1 #0891b2);
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #0891b2, stop:1 #0e7490);
            }
        """)
        self.backup_categories_btn.clicked.connect(self.backup_categories)

        self.backup_accounts_btn = QPushButton("💾 Hesap Kategorilerini Yedekle")
        self.backup_accounts_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #8b5cf6, stop:1 #7c3aed);
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #7c3aed, stop:1 #6d28d9);
            }
        """)
        self.backup_accounts_btn.clicked.connect(self.backup_account_categories)

        self.full_backup_btn = QPushButton("💾 Tam Yedek Oluştur")
        self.full_backup_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #10b981, stop:1 #059669);
                font-size: 16px;
                padding: 16px 32px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #059669, stop:1 #047857);
            }
        """)
        self.full_backup_btn.clicked.connect(self.create_full_backup)

        backup_layout.addWidget(self.backup_categories_btn, 0, 0)
        backup_layout.addWidget(self.backup_accounts_btn, 0, 1)
        backup_layout.addWidget(self.full_backup_btn, 1, 0, 1, 2)
        backup_frame.setLayout(backup_layout)

        layout.addWidget(backup_frame)
        layout.addStretch()

        widget.setLayout(layout)
        return widget

    # Diğer metodlar (select_batch_files, import_from_excel, etc.) aynı kalabilir
    # Sadece tasarım güncellendi, fonksiyonalite korundu

    def select_batch_files(self):
        """Toplu içe aktarma için dosyaları seç"""
        files, _ = QFileDialog.getOpenFileNames(
            self, "İçe Aktarılacak Dosyaları Seç",
            "", "Desteklenen Dosyalar (*.json *.txt *.csv);;Tüm Dosyalar (*)"
        )

        if files:
            self.file_list.clear()
            for file_path in files:
                item = QListWidgetItem(f"📄 {os.path.basename(file_path)}")
                item.setData(Qt.UserRole, file_path)
                item.setToolTip(file_path)
                self.file_list.addItem(item)

    def clear_file_list(self):
        """Dosya listesini temizle"""
        self.file_list.clear()

    def select_excel_file(self):
        """Excel dosyası seç"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Excel Dosyası Seç",
            "", "Excel Dosyaları (*.xlsx *.xls)"
        )

        if file_path:
            self.excel_file_path = file_path
            self.excel_file_label.setText(f"📄 {os.path.basename(file_path)}")
            self.excel_file_label.setStyleSheet("color: #059669; font-weight: 600;")
            self.import_excel_btn.setEnabled(True)

    def select_validation_files(self):
        """Doğrulama için dosyaları seç"""
        files, _ = QFileDialog.getOpenFileNames(
            self, "Doğrulanacak Dosyaları Seç",
            "", "Desteklenen Dosyalar (*.json *.txt *.csv);;Tüm Dosyalar (*)"
        )

        if files:
            self.validation_file_list.clear()
            for file_path in files:
                item = QListWidgetItem(f"📄 {os.path.basename(file_path)}")
                item.setData(Qt.UserRole, file_path)
                item.setToolTip(file_path)
                self.validation_file_list.addItem(item)

    def start_batch_import(self):
        """Toplu içe aktarmayı başlat"""
        if self.file_list.count() == 0:
            QMessageBox.warning(self, "⚠️ Uyarı", "Lütfen önce dosyaları seçin!")
            return

        file_paths = []
        for i in range(self.file_list.count()):
            item = self.file_list.item(i)
            file_paths.append(item.data(Qt.UserRole))

        options = {
            'skip_duplicates': self.skip_duplicates_check.isChecked(),
            'create_backup': self.create_backup_check.isChecked()
        }

        self.start_worker("batch_import", file_paths, options)

    def import_from_excel(self):
        """Excel'den içe aktarma"""
        if not hasattr(self, 'excel_file_path'):
            QMessageBox.warning(self, "⚠️ Uyarı", "Lütfen önce Excel dosyasını seçin!")
            return

        self.start_worker("excel_import", [self.excel_file_path], {})

    def start_validation(self):
        """Doğrulamayı başlat"""
        if self.validation_file_list.count() == 0:
            QMessageBox.warning(self, "⚠️ Uyarı", "Lütfen önce dosyaları seçin!")
            return

        file_paths = []
        for i in range(self.validation_file_list.count()):
            item = self.validation_file_list.item(i)
            file_paths.append(item.data(Qt.UserRole))

        self.start_worker("validate_data", file_paths, {})

    def backup_categories(self):
        """Kategorileri yedekle"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Kategori Yedeği Kaydet", 
                f"kategori_yedek_{timestamp}.json",
                "JSON Dosyaları (*.json)"
            )

            if file_path:
                if mysql_manager:
                    categories = mysql_manager.get_categories('icerik')
                    backup_data = {
                        'backup_date': datetime.now().isoformat(),
                        'backup_type': 'categories',
                        'categories': categories
                    }

                    with open(file_path, 'w', encoding='utf-8') as f:
                        json.dump(backup_data, f, ensure_ascii=False, indent=2)

                    QMessageBox.information(self, "✅ Başarılı", f"Kategoriler yedeklendi:\n{file_path}")
                else:
                    QMessageBox.warning(self, "❌ Hata", "MySQL bağlantısı bulunamadı")

        except Exception as e:
            QMessageBox.critical(self, "❌ Hata", f"Yedekleme hatası: {str(e)}")

    def backup_account_categories(self):
        """Hesap kategorilerini yedekle"""
        QMessageBox.information(self, "💡 Bilgi", "Hesap kategorileri yedeklemesi henüz hazırlanıyor...")

    def create_full_backup(self):
        """Tam yedek oluştur"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Tam Yedek Dosyasını Kaydet",
                f"aktweetor_tam_yedek_{timestamp}.json",
                "JSON Dosyaları (*.json)"
            )

            if file_path:
                # Tüm kategori verilerini topla
                categories = mysql_manager.get_categories('icerik')

                # Tüm hesap kategorilerini topla
                all_account_categories = []

                # Giriş yapılan hesaplar
                for acc in user_manager.get_all_users():
                    acc_cats = mysql_manager.get_account_categories(acc['kullanici_adi'], 'giris_yapilan')
                    for cat in acc_cats:
                        cat['kullanici_adi'] = acc['kullanici_adi']
                        cat['hesap_turu'] = 'giris_yapilan'
                        all_account_categories.append(cat)

                # Hedef hesaplar
                for acc in mysql_manager.get_all_targets():
                    acc_cats = mysql_manager.get_account_categories(acc['kullanici_adi'], 'hedef')
                    for cat in acc_cats:
                        cat['kullanici_adi'] = acc['kullanici_adi']
                        cat['hesap_turu'] = 'hedef'
                        all_account_categories.append(cat)

                backup_data = {
                    'backup_date': datetime.now().isoformat(),
                    'version': '2.0',
                    'categories': categories,
                    'account_categories': all_account_categories
                }

                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(backup_data, f, ensure_ascii=False, indent=2)

                self.log_text.append(f"✅ Tam yedek oluşturuldu: {file_path}")
                QMessageBox.information(self, "✅ Başarılı", "Tam yedek başarıyla oluşturuldu!")

        except Exception as e:
            QMessageBox.critical(self, "❌ Hata", f"Yedek oluşturulamadı: {str(e)}")

    def download_excel_template(self):
        """Excel şablonunu indir"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            file_path, _ = QFileDialog.getSaveFileName(
                self, "Excel Şablonunu Kaydet",
                "aktweetor_template.xlsx", "Excel Dosyaları (*.xlsx)"
            )

            if file_path:
                wb = openpyxl.Workbook()

                # Kategoriler sayfası
                ws_categories = wb.active
                ws_categories.title = "Kategoriler"

                headers = ['Ana Kategori', 'Alt Kategori', 'Açıklama']
                for col, header in enumerate(headers, 1):
                    cell = ws_categories.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')

                # Örnek veriler
                example_data = [
                    ['Siyasi Eğilim', 'Muhafazakar', 'Muhafazakar görüşlü içerikler'],
                    ['Siyasi Eğilim', 'Liberal', 'Liberal görüşlü içerikler'],
                    ['Fotoğraf İçeriği', 'Parti Logosu', 'Siyasi parti logoları']
                ]

                for row, data in enumerate(example_data, 2):
                    for col, value in enumerate(data, 1):
                        ws_categories.cell(row=row, column=col, value=value)

                # Hesap kategorileri sayfası
                ws_accounts = wb.create_sheet("Hesap Kategorileri")

                headers = ['Kullanıcı Adı', 'Hesap Türü', 'Ana Kategori', 'Alt Kategori', 'Kategori Değeri']
                for col, header in enumerate(headers, 1):
                    cell = ws_accounts.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')

                # Örnek veriler
                example_data = [
                    ['ornekkullanici1', 'Giriş Yapılan', 'Yaş Grubu', '', 'Genç (18-30)'],
                    ['ornekkullanici2', 'Hedef', 'Cinsiyet', '', 'Erkek'],
                    ['ornekkullanici3', 'Hedef', 'Siyasi Eğilim', 'Muhafazakar', 'Seçili']
                ]

                for row, data in enumerate(example_data, 2):
                    for col, value in enumerate(data, 1):
                        ws_accounts.cell(row=row, column=col, value=value)

                # Sütun genişliklerini ayarla
                for ws in wb.worksheets:
                    for column_cells in ws.columns:
                        length = max(len(str(cell.value or '')) for cell in column_cells)
                        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 30)

                wb.save(file_path)

                QMessageBox.information(self, "✅ Başarılı", f"Excel şablonu oluşturuldu:\n{file_path}")

        except ImportError:
            QMessageBox.warning(self, "⚠️ Uyarı", "Excel şablonu için openpyxl kütüphanesi gerekli.\nLütfen 'pip install openpyxl' çalıştırın.")
        except Exception as e:
            QMessageBox.critical(self, "❌ Hata", f"Şablon oluşturulamadı: {str(e)}")

    def start_worker(self, operation_type, file_paths, options):
        """Worker thread'i başlat"""
        self.worker_thread = AdvancedImportThread(operation_type, file_paths, options)
        self.worker_thread.progress.connect(self.progress_bar.setValue)
        self.worker_thread.status.connect(self.status_label.setText)
        self.worker_thread.log.connect(self.log_text.append)
        self.worker_thread.finished.connect(self.on_operation_finished)
        self.worker_thread.error.connect(self.on_operation_error)

        # UI'yi güncelle
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)

        self.worker_thread.start()

    def on_operation_finished(self, message):
        """İşlem tamamlandığında"""
        self.progress_bar.setVisible(False)
        self.status_label.setText("✅ " + message)
        self.log_text.append("✅ " + message)

        QMessageBox.information(self, "✅ Başarılı", message)

    def on_operation_error(self, error_message):
        """İşlem hatası"""
        self.progress_bar.setVisible(False)
        self.status_label.setText("❌ " + error_message)
        self.log_text.append("❌ " + error_message)

        QMessageBox.critical(self, "❌ Hata", error_message)

    def closeEvent(self, event):
        """Dialog kapatılırken"""
        if self.worker_thread and self.worker_thread.isRunning():
            self.worker_thread.terminate()
            self.worker_thread.wait()
        event.accept()
    
    def setup_import_tab(self):
        """İçe aktarma sekmesi"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(20)

        # Bilgi kartı
        info_card = QFrame()
        info_card.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #dbeafe, stop:1 #bfdbfe);
                border: 2px solid #93c5fd;
                border-radius: 16px;
                padding: 20px;
            }
        """)

        info_layout = QVBoxLayout()

        info_title = QLabel("📁 Kategori İçe Aktarma")
        info_title.setStyleSheet("font-size: 20px; font-weight: 700; color: #1e40af; margin-bottom: 8px;")

        info_desc = QLabel("Kategori ve hesap verilerini içe aktarın\n• JSON ve TXT formatlarını destekler\n• Hızlı ve kolay içe aktarma")
        info_desc.setStyleSheet("color: #3730a3; font-size: 14px; line-height: 1.4;")

        info_layout.addWidget(info_title)
        info_layout.addWidget(info_desc)
        info_card.setLayout(info_layout)

        layout.addWidget(info_card)

        # Dosya seçimi
        file_frame = QGroupBox("📁 Dosya Seçimi")
        file_layout = QVBoxLayout()

        file_buttons_layout = QHBoxLayout()

        self.import_categories_btn = QPushButton("📁 Kategori Dosyası Seç (.txt)")
        self.import_categories_btn.clicked.connect(self.import_categories)

        self.import_json_btn = QPushButton("📄 JSON Kategorileri Seç")
        self.import_json_btn.clicked.connect(self.import_json_categories)
        
        self.import_account_categories_btn = QPushButton("👤 Hesap Kategorileri Dosyası Seç")
        self.import_account_categories_btn.clicked.connect(self.import_account_categories)

        file_buttons_layout.addWidget(self.import_categories_btn)
        file_buttons_layout.addWidget(self.import_json_btn)
        file_buttons_layout.addWidget(self.import_account_categories_btn)
        file_buttons_layout.addStretch()

        self.file_list = QListWidget()
        self.file_list.setMinimumHeight(150)

        file_layout.addLayout(file_buttons_layout)
        file_layout.addWidget(self.file_list)
        file_frame.setLayout(file_layout)

        layout.addWidget(file_frame)

        layout.addStretch()

        widget.setLayout(layout)
        return widget

    def import_json_categories(self):
        """JSON kategori dosyası içe aktar"""
        file_path, _ = QFileDialog.getOpenFileName(self, "JSON Kategori Dosyası Seç", "", "JSON Dosyaları (*.json)")
        if file_path:
            count = mysql_manager.import_categories_from_json(file_path)
            self.log_text.append(f"✅ {count} JSON kategorisi içe aktarıldı")
            # Kategori listelerini yenile
            if hasattr(self.parent(), 'load_photo_content_categories'):
                self.parent().load_photo_content_categories()
            if hasattr(self.parent(), 'load_profile_content_categories'):
                self.parent().load_profile_content_categories()

    def import_account_categories(self):
        """Hesap kategorileri dosyası içe aktar"""
        file_path, _ = QFileDialog.getOpenFileName(self, "Hesap Kategorileri Dosyası Seç", "", "Metin Dosyaları (*.txt)")
        if file_path:
            # Hesap türü seçimi gerekli - şimdilik hedef hesap olarak varsayalım
            count = mysql_manager.import_account_categories_from_file(file_path, 'hedef')
            self.log_text.append(f"✅ {count} hesap kategorisi içe aktarıldı")

    def import_categories(self):
        """Kategori dosyası içe aktar"""
        file_path, _ = QFileDialog.getOpenFileName(self, "Kategori Dosyası Seç", "", "Metin Dosyaları (*.txt)")
        if file_path:
            count = mysql_manager.import_categories_from_file(file_path)
            self.log_text.append(f"✅ {count} kategori içe aktarıldı")
            # Kategori listelerini yenile
            if hasattr(self.parent(), 'load_photo_content_categories'):
                self.parent().load_photo_content_categories()
            if hasattr(self.parent(), 'load_profile_content_categories'):
                self.parent().load_profile_content_categories()